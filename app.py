# app.py
import streamlit as st
import pandas as pd
import openpyxl
import io
import requests
from datetime import datetime

st.set_page_config(page_title="Medical Quotation Generator", layout="wide")

# ------------------------------------------------------------
# LOGIN
# ------------------------------------------------------------
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if not st.session_state.logged_in:
    st.title("Login Required")
    username = st.text_input("Username")
    password = st.text_input("Password", type="password")

    if st.button("Login"):
        if username == "admin" and password == "Jamela2003":
            st.session_state.logged_in = True
            st.success("Login successful! Reload or interact with the app to continue.")
        else:
            st.error("Invalid credentials")
    st.stop()

# ------------------------------------------------------------
# CONFIG
# ------------------------------------------------------------
COMPONENT_KEYS = {"PELVIS", "CONSUMABLES", "FF", "IV", "IV CONTRAST", "IV CONTRAST 100MLS"}
GARBAGE_KEYS = {"TOTAL", "CO-PAYMENT", "CO PAYMENT", "CO - PAYMENT", "CO", ""}
MAIN_CATEGORIES = set()

# ------------------------------------------------------------
# HELPERS
# ------------------------------------------------------------
def clean_text(x):
    if pd.isna(x):
        return ""
    return str(x).replace("\xa0", " ").strip()

def safe_int(x, default=1):
    try:
        return int(float(str(x).replace(",", "").strip()))
    except:
        return default

def safe_float(x, default=0.0):
    try:
        return float(str(x).replace(",", "").strip())
    except:
        return default

# ------------------------------------------------------------
# PARSER
# ------------------------------------------------------------
def load_charge_sheet(file):
    df_raw = pd.read_excel(file, header=None, dtype=object, engine="openpyxl")
    while df_raw.shape[1] < 5:
        df_raw[df_raw.shape[1]] = None
    df_raw = df_raw.iloc[:, :5]
    df_raw.columns = ["A_EXAM", "B_TARIFF", "C_MOD", "D_QTY", "E_AMOUNT"]

    structured = []
    current_category = None
    current_subcategory = None

    for _, r in df_raw.iterrows():
        exam = clean_text(r["A_EXAM"])
        if not exam:
            continue
        exam_u = exam.upper().strip()

        if exam_u in MAIN_CATEGORIES or exam_u.endswith("SCAN") or exam_u in {"XRAY", "MRI", "ULTRASOUND"}:
            MAIN_CATEGORIES.add(exam_u)
            current_category = exam
            current_subcategory = None
            continue
        if exam_u in GARBAGE_KEYS:
            continue
        if clean_text(r["B_TARIFF"]) == "" and clean_text(r["E_AMOUNT"]) == "":
            current_subcategory = exam
            continue
        if not current_category:
            continue

        structured.append({
            "CATEGORY": current_category,
            "SUBCATEGORY": current_subcategory,
            "SCAN": exam,
            "IS_MAIN_SCAN": exam_u not in COMPONENT_KEYS,
            "TARIFF": safe_float(r["B_TARIFF"], None),
            "MODIFIER": clean_text(r["C_MOD"]),
            "QTY": safe_int(r["D_QTY"], 1),
            "AMOUNT": safe_float(r["E_AMOUNT"], 0.0)
        })

    return pd.DataFrame(structured)

# ------------------------------------------------------------
# EXCEL HELPERS
# ------------------------------------------------------------
def write_safe(ws, r, c, value):
    if not c:
        return
    try:
        ws.cell(row=r, column=c).value = value
    except:
        for mr in ws.merged_cells.ranges:
            if ws.cell(row=r, column=c).coordinate in mr:
                ws.cell(row=mr.min_row, column=mr.min_col).value = value
                return

def append_after_label(ws, r, c, value):
    if not value:
        return
    cell = ws.cell(row=r, column=c)
    existing = str(cell.value) if cell.value else ""
    cell.value = f"{existing.strip()} {value}".strip()

def write_below_label(ws, r, c, value):
    target = ws.cell(row=r + 1, column=c)
    try:
        target.value = value
    except:
        for mr in ws.merged_cells.ranges:
            if target.coordinate in mr:
                ws.cell(row=mr.min_row, column=mr.min_col).value = value
                return

def find_template_positions(ws):
    pos = {}
    headers = ["DESCRIPTION", "TARIFF", "TARRIF", "MOD", "QTY", "FEES", "AMOUNT"]

    for row in ws.iter_rows(min_row=1, max_row=200):
        for cell in row:
            if not cell.value:
                continue
            t = str(cell.value).upper()
            if "PATIENT" in t:
                pos["patient_cell"] = (cell.row, cell.column)
            if "MEMBER" in t:
                pos["member_cell"] = (cell.row, cell.column)
            if "PROVIDER" in t or "MEDICAL AID" in t:
                pos["provider_cell"] = (cell.row, cell.column)
            if t.strip() == "DATE":
                pos["date_cell"] = (cell.row, cell.column)
            if any(h in t for h in headers):
                pos.setdefault("cols", {})
                pos["table_start_row"] = cell.row + 1
                for h in headers:
                    if h in t:
                        pos["cols"][h] = cell.column
    return pos

def fill_excel_template(template_file, patient, member, provider, scan_rows, date_value=None, full_scan_name=""):
    wb = openpyxl.load_workbook(template_file)
    ws = wb.active
    pos = find_template_positions(ws)

    if "patient_cell" in pos:
        append_after_label(ws, *pos["patient_cell"], patient)
    if "member_cell" in pos:
        append_after_label(ws, *pos["member_cell"], member)
    if "provider_cell" in pos:
        append_after_label(ws, *pos["provider_cell"], provider)
    if "date_cell" in pos and date_value:
        write_below_label(ws, *pos["date_cell"], date_value.strftime("%d/%m/%Y"))

    description_col = pos["cols"].get("DESCRIPTION")
    if not description_col:
        st.error("Could not find DESCRIPTION column in template. Check the header.")
        return None

    rowptr = pos.get("table_start_row", 22)
    grand_total = 0.0

    # Debug log
    st.write("DEBUG: Starting Excel rowptr =", rowptr)
    st.write("DEBUG: Number of selected scans =", len(scan_rows))
    st.write("DEBUG: Full scan name =", full_scan_name)

    # Write selected scans
    for sr in scan_rows:
        is_main = sr.get("IS_MAIN_SCAN", True)
        scan_desc = sr.get("FINAL_SCAN") or sr.get("SCAN", "")
        tariff = sr.get("TARIFF", 0.0)
        modifier = sr.get("MODIFIER", "")
        qty = sr.get("QTY", 1)
        amount = sr.get("AMOUNT", 0.0)

        write_safe(
            ws,
            rowptr,
            description_col,
            scan_desc if is_main else "   " + scan_desc
        )
        write_safe(
            ws,
            rowptr,
            pos["cols"].get("TARIFF") or pos["cols"].get("TARRIF"),
            tariff
        )
        write_safe(ws, rowptr, pos["cols"].get("MOD"), modifier)
        write_safe(ws, rowptr, pos["cols"].get("QTY"), qty)
        write_safe(ws, rowptr, pos["cols"].get("FEES"), round(amount, 2))
        grand_total += amount
        rowptr += 1
        st.write(f"DEBUG: Wrote scan '{scan_desc}' at row {rowptr-1}")

    # Add full scan name as separate row if provided
    if full_scan_name.strip():
        write_safe(ws, rowptr, description_col, full_scan_name.strip())
        write_safe(ws, rowptr, pos["cols"].get("TARIFF") or pos["cols"].get("TARRIF"), 0)
        write_safe(ws, rowptr, pos["cols"].get("MOD"), "")
        write_safe(ws, rowptr, pos["cols"].get("QTY"), 1)
        write_safe(ws, rowptr, pos["cols"].get("FEES"), 0.0)
        st.write(f"DEBUG: Wrote full scan '{full_scan_name.strip()}' at row {rowptr}")
        rowptr += 1

    write_safe(ws, 22, pos["cols"].get("AMOUNT"), round(grand_total, 2))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ------------------------------------------------------------
# LOAD EXTERNAL FILES
# ------------------------------------------------------------
@st.cache_data(show_spinner=False)
def fetch_charge_sheet():
    url = (
        "https://docs.google.com/spreadsheets/d/e/"
        "2PACX-1vTmaRisOdFHXmFsxVA7Fx0odUq1t2QfjMvRBKqeQPgoJUdrIgSU6UhNs_-dk4jfVQ"
        "/pub?output=xlsx"
    )
    try:
        return load_charge_sheet(url)
    except Exception as e:
        st.error(f"Failed to load charge sheet: {e}")
        return pd.DataFrame()

@st.cache_data(show_spinner=False)
def fetch_quote_template():
    url = "https://www.dropbox.com/scl/fi/iup7nwuvt5y74iu6dndak/new-template.xlsx?rlkey=5oleriordmi3bktx2f8kx36ew&st=oxytxpph&dl=1"
    try:
        response = requests.get(url, allow_redirects=True, timeout=30)
        response.raise_for_status()
        content = response.content
        if not content.startswith(b"PK"):
            raise ValueError("Downloaded template is not a valid Excel (.xlsx) file.")
        return io.BytesIO(content)
    except Exception as e:
        st.error(f"Failed to fetch template: {e}")
        return None

# ------------------------------------------------------------
# STREAMLIT UI
# ------------------------------------------------------------
st.title("Medical Quotation Generator")

patient = st.text_input("Patient Name")
member = st.text_input("Medical Aid / Member Number")
provider = st.text_input("Medical Aid Provider", value="CIMAS")
quotation_date = st.date_input("Quotation Date", value=datetime.today())

# Full scan name input
full_scan_name = st.text_input("Full Scan Name for Details List", value="")

# Load charge sheet
if "df" not in st.session_state:
    st.session_state.df = fetch_charge_sheet()
    st.success("Charge sheet loaded automatically!")

df = st.session_state.df
if df.empty:
    st.stop()

main_sel = st.selectbox(
    "Select Main Category",
    sorted(df["CATEGORY"].dropna().unique())
)

subcats = sorted(df[df["CATEGORY"] == main_sel]["SUBCATEGORY"].dropna().unique())
sub_sel = st.selectbox("Select Subcategory", subcats) if subcats else None

scans = (
    df[(df["CATEGORY"] == main_sel) & (df["SUBCATEGORY"] == sub_sel)]
    if sub_sel else df[df["CATEGORY"] == main_sel]
).reset_index(drop=True)

scans["label"] = scans.apply(lambda r: f"{r['SCAN']} | Amount {r['AMOUNT']}", axis=1)

selected = st.multiselect(
    "Select scans to include",
    options=scans.index.tolist(),
    format_func=lambda i: scans.at[i, "label"]
)

selected_rows = [scans.iloc[i].to_dict() for i in selected]

if selected_rows:
    edits_df = pd.DataFrame(selected_rows)

    if "FINAL_SCAN" not in edits_df.columns:
        edits_df["FINAL_SCAN"] = edits_df["SCAN"]

    st.subheader("Edit and Preview Final Descriptions")
    edited_df = st.data_editor(
        edits_df,
        column_config={
            "FINAL_SCAN": st.column_config.TextColumn("Final Description", max_chars=100),
            "MODIFIER": st.column_config.TextColumn("Modifier", disabled=True),
            "AMOUNT": st.column_config.NumberColumn("Amount", format="$%.2f", disabled=True),
        },
        use_container_width=True
    )

    edited_df = edited_df.fillna({
        "IS_MAIN_SCAN": True,
        "TARIFF": 0.0,
        "MODIFIER": "",
        "QTY": 1,
        "CATEGORY": "",
        "SUBCATEGORY": "",
        "AMOUNT": 0.0
    })

    selected_rows = edited_df.to_dict("records")

    total_amount = sum(r["AMOUNT"] for r in selected_rows)
    st.metric("Grand Total", f"${total_amount:,.2f}")

    if st.button("Generate & Download Quotation"):
        template_file = fetch_quote_template()
        if template_file:
            out = fill_excel_template(
                template_file, patient, member, provider, selected_rows,
                date_value=quotation_date,
                full_scan_name=full_scan_name
            )
            if out:
                st.download_button(
                    "Download Quotation",
                    data=out,
                    file_name="quotation.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
